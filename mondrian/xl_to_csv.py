# This script can be run to transform .xls files to .csv files
# Override the inpu_dir variable with the local folder that contain the .xls files.
import openpyxl.utils
import pandas as pd
import os
import csv
from openpyxl.worksheet.worksheet import Worksheet

print("Excel to csv script")

input_dir = "./in"
output_dir = "./out"

with open(output_dir + "stats_hidden.csv", "w") as f:
    csv_writer = csv.writer(f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
    csv_writer.writerow(["filename", "hidden_rows", "hidden_cols"])

with open(output_dir + "stats_merged.csv", "w") as f:
    csv_writer = csv.writer(f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
    csv_writer.writerow(["filename", "merged_area_min_col", "merged_area_min_row", "merged_area_max_col", "merged_area_max_row"])

with open(output_dir + "stats_corrupt.txt", "w") as f:
    f.write("")

xls_files = sorted(os.listdir(input_dir))

for file in xls_files:
    if file[-3:] == "lsx":
        print(file)
        xls = pd.ExcelFile(input_dir + file)

        try:
            xls = openpyxl.load_workbook(input_dir + file)
        except Exception as e:
            print("\tParsing error! ", e)
            with open(output_dir + "stats_corrupt.txt", "a") as f:
                f.write(file)
                f.write("\n\tparse exception: " + str(e))
                f.write("\n")
            continue
        for sheet in xls.sheetnames:
            ws = xls[sheet]
            if ws.sheet_state != 'hidden' and isinstance(ws, Worksheet):

                hidden_cols = []
                hidden_rows = []
                for letter, col_dimension in ws.column_dimensions.items():
                    if col_dimension.hidden:
                        index = openpyxl.utils.column_index_from_string(letter) - 1
                        hidden_cols.append(index)

                for index, row_dimension in ws.row_dimensions.items():
                    if row_dimension.hidden:
                        hidden_rows.append(index - 1)

                data_xls = pd.read_excel(input_dir + file, sheet, index_col=None, header=None)

                if not data_xls.empty:
                    print("\t" + sheet)

                    # fill range cells
                    for cell_range in ws.merged_cells.ranges:
                        min_col, min_row, max_col, max_row = [b - 1 for b in cell_range.bounds]  # offset by 1
                        fill = data_xls.iloc[min_row, min_col]
                        data_xls.loc[min_row:max_row, min_col:max_col] = fill

                    hidden_rows = list(set(range(len(data_xls))) & set(hidden_rows))  # exclude empty lines
                    unhidden_cols = list(set(data_xls.columns) - set(hidden_cols))
                    data_xls = data_xls[unhidden_cols]
                    data_xls.drop(hidden_rows, inplace=True)

                    fname = file + "_" + sheet + ".csv"
                    data_xls.to_csv(output_dir + fname, encoding='utf-8', index=False, header=None)

                    # output if it had merged or hidden rows/columns
                    if hidden_rows != [] or hidden_cols != []:
                        with open(output_dir + "stats_hidden.csv", "a") as f:
                            csv_writer = csv.writer(f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                            csv_writer.writerow([fname, hidden_rows, hidden_cols])

                    if ws.merged_cells.ranges:
                        with open(output_dir + "stats_merged.csv", "a") as f:
                            csv_writer = csv.writer(f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                            for r in ws.merged_cells.ranges:
                                csv_writer.writerow([fname, r.min_col - 1, r.min_row - 1, r.max_col - 1, r.max_row - 1])
